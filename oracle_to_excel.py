import cx_Oracle as ora
import openpyxl as op
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re

ip = ''
port = 
SERVER = 
SERVICE_NAME = 
dsn_tns = ora.makedsn(ip, port, SERVER, SERVICE_NAME)

connection = ora.connect(user = "", password = "", dsn = dsn_tns)

wb = op.Workbook()
ws = wb.active

cursor = connection.cursor()
cursor.execute(
     "ALTER SESSION SET NLS_DATE_FORMAT = 'DD.MM.YY'"
     " NLS_TIMESTAMP_FORMAT = 'DD.MM.YY'")

command = "select pararms from tablename where report_dt >= '28.12.21'"
print(command)
cursor.execute(command)


for row_n in cursor.description:
    ws.cell(row = 1, column = cursor.description.index(row_n)+1).value = row_n[0]


n = 2
for row_n in cursor:
    for col_n in range(len(row_n)):
        try:
            ws.cell(row = n, column = col_n+1).value = row_n[col_n]
        except op.utils.exceptions.IllegalCharacterError:
            ws.cell(row = n, column = col_n+1).value = ILLEGAL_CHARACTERS_RE.sub("", row_n[col_n])
    n+=1
    if n%10000 == 0:
        print(n)
        
wb.save(filepath)
